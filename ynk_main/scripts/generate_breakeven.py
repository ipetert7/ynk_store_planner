"""Script para generar análisis de breakeven por sucursal en formato Excel."""
from __future__ import annotations

import argparse
import sys
from pathlib import Path

# Agregar el directorio raíz del proyecto al path
project_root = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(project_root / "src"))

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from ynk_modelo.config import PROJECT_ROOT
from ynk_modelo.domain.eerr import build_breakeven_table_full_range


def format_excel_file(file_path: Path) -> None:
    """Formatea el archivo Excel con estilos apropiados."""
    wb = load_workbook(file_path)
    ws = wb.active
    
    # Estilo para el encabezado
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Formatear encabezado
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Ajustar ancho de columnas
    ws.column_dimensions["A"].width = 25  # Sucursal
    ws.column_dimensions["B"].width = 22  # Margen contribución (%)
    ws.column_dimensions["C"].width = 30  # Venta necesaria para breakeven ($)
    
    # Aplicar formatos de número a columnas completas (más eficiente)
    # Columna B: Margen contribución (porcentaje)
    for row in range(2, min(ws.max_row + 1, 10002)):  # Limitar a primeras 10k filas para rendimiento
        cell = ws[f"B{row}"]
        if cell.value is not None:
            cell.number_format = "0.0"
    
    # Columna C: Venta necesaria (moneda)
    for row in range(2, min(ws.max_row + 1, 10002)):  # Limitar a primeras 10k filas para rendimiento
        cell = ws[f"C{row}"]
        if cell.value is not None and not (isinstance(cell.value, str) and cell.value == "-"):
            cell.number_format = "#,##0"
    
    # Congelar primera fila
    ws.freeze_panes = "A2"
    
    wb.save(file_path)


def generate_breakeven_excel(
    output_path: Path | None = None,
    margen_paso: float = 0.1,
) -> Path:
    """Genera archivo Excel con análisis de breakeven por sucursal.
    
    Args:
        output_path: Ruta del archivo Excel de salida. Si es None, usa output/breakeven_analysis.xlsx
        margen_paso: Paso del margen de contribución en porcentaje (default: 0.1%)
    
    Returns:
        Path del archivo generado
    """
    if output_path is None:
        output_dir = PROJECT_ROOT / "output"
        output_dir.mkdir(exist_ok=True)
        output_path = output_dir / "breakeven_analysis.xlsx"
    else:
        output_path = Path(output_path)
        output_path.parent.mkdir(parents=True, exist_ok=True)
    
    print(f"Calculando breakeven para todas las sucursales (paso: {margen_paso}%)...")
    print("Nota: El cálculo se realiza sin aplicar el factor de diciembre (como en el simulador HTML).")
    df = build_breakeven_table_full_range(margen_paso=margen_paso, usar_factor_diciembre=False)
    
    # Renombrar columnas al español
    df_export = df.rename(columns={
        "Sucursal": "Sucursal",
        "Margen_contribucion": "Margen contribución (%)",
        "Venta_necesaria": "Venta necesaria para breakeven ($)",
    })
    
    # Ordenar por Sucursal y Margen contribución
    df_export = df_export.sort_values(["Sucursal", "Margen contribución (%)"])
    
    # Reemplazar NaN por "-" para mejor visualización
    df_export["Venta necesaria para breakeven ($)"] = df_export["Venta necesaria para breakeven ($)"].fillna("-")
    
    print(f"Generando archivo Excel: {output_path}")
    print("Escribiendo datos a Excel (esto puede tomar varios minutos para archivos grandes)...")
    
    # Escribir archivo temporal primero para evitar problemas de I/O
    temp_path = output_path.with_suffix(".tmp.xlsx")
    
    try:
        # Escribir sin formateo primero para mejor rendimiento
        with pd.ExcelWriter(temp_path, engine="openpyxl", mode="w") as writer:
            df_export.to_excel(writer, index=False, sheet_name="Breakeven")
        
        # Mover archivo temporal a destino final
        temp_path.replace(output_path)
        print("Archivo base generado exitosamente.")
        
        # Intentar formateo básico solo en encabezado (rápido)
        try:
            wb = load_workbook(output_path)
            ws = wb.active
            
            # Formatear solo encabezado
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF", size=11)
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Ajustar ancho de columnas
            ws.column_dimensions["A"].width = 25
            ws.column_dimensions["B"].width = 22
            ws.column_dimensions["C"].width = 30
            
            # Congelar primera fila
            ws.freeze_panes = "A2"
            
            wb.save(output_path)
            print("Formato básico aplicado (encabezado y anchos de columna).")
        except Exception as e:
            print(f"Advertencia: No se pudo aplicar formato adicional: {e}")
            print("El archivo se generó correctamente pero sin formato avanzado.")
            
    except Exception as e:
        # Si hay error, limpiar archivo temporal
        if temp_path.exists():
            temp_path.unlink()
        raise
    
    print(f"Archivo generado exitosamente: {output_path}")
    print(f"Total de filas: {len(df_export)}")
    print(f"Total de sucursales: {df_export['Sucursal'].nunique()}")
    
    return output_path


def main() -> None:
    """Función principal del script."""
    parser = argparse.ArgumentParser(
        description="Genera análisis de breakeven por sucursal en formato Excel."
    )
    parser.add_argument(
        "--output",
        type=Path,
        default=None,
        help="Ruta del archivo Excel de salida (por defecto: output/breakeven_analysis.xlsx)",
    )
    parser.add_argument(
        "--paso",
        type=float,
        default=0.1,
        help="Paso del margen de contribución en porcentaje (por defecto: 0.1%%)",
    )
    
    args = parser.parse_args()
    
    try:
        generate_breakeven_excel(output_path=args.output, margen_paso=args.paso)
    except Exception as e:
        print(f"Error al generar el archivo: {e}")
        raise


if __name__ == "__main__":
    main()

